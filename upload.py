import argparse
import openpyxl

parser = argparse.ArgumentParser()
parser.add_argument("-f","--file",required=True)
args = parser.parse_args()


wb = openpyxl.load_workbook(args.file,data_only=True)
ws = wb.active

r=2

while 1:
	if ws.cell(column=1,row=r).value is None:
		break
	else:
		for c in range(1,5):
			print(ws.cell(column=c,row=r).value)
		r=r+1
